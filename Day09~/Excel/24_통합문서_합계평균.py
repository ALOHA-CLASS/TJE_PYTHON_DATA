#!/usr/bin/env python3
import glob
import os
import sys
from openpyxl import Workbook, load_workbook

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력폴더, 출력파일
input_folder = path + '/input/'
output_file = path + '/output/' + input('출력 파일 : ')

output_workbook = Workbook()
output_worksheet = output_workbook.active
output_worksheet.title = 'sums_and_averages'

all_data = []
sales_column_index = 3

header = ['workbook', 'worksheet', 'worksheet_total', 'worksheet_average',\
 					'workbook_total', 'workbook_average']
all_data.append(header)

for input_file in glob.glob(os.path.join(input_folder, '*.xlsx')):
    workbook = load_workbook(filename=input_file, read_only=True)
    list_of_totals = []
    list_of_numbers = []
    workbook_output = []
    for worksheet in workbook.worksheets:
        total_sales = 0
        number_of_sales = 0
        worksheet_list = []
        worksheet_list.append(os.path.basename(input_file))
        worksheet_list.append(worksheet.title)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            try:
                total_sales += float(str(row[sales_column_index]).strip('$').replace(',', ''))
                number_of_sales += 1
            except ValueError:
                pass
        average_sales = '%.2f' % (total_sales / number_of_sales) if number_of_sales > 0 else '0.00'
        worksheet_list.append(total_sales)
        worksheet_list.append(float(average_sales))
        list_of_totals.append(total_sales)
        list_of_numbers.append(float(number_of_sales))
        workbook_output.append(worksheet_list)
    workbook_total = sum(list_of_totals)
    workbook_average = sum(list_of_totals) / sum(list_of_numbers) if sum(list_of_numbers) > 0 else 0.0
    for list_element in workbook_output:
        list_element.append(workbook_total)
        list_element.append(workbook_average)
    all_data.extend(workbook_output)

for list_index, output_list in enumerate(all_data):
    for element_index, element in enumerate(output_list):
        output_worksheet.cell(row=list_index + 1, column=element_index + 1, value=element)

output_workbook.save(output_file)
