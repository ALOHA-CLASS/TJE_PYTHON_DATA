#!/usr/bin/env python3
import sys
from datetime import date
from openpyxl import Workbook, load_workbook

import os

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력파일, 출력파일
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

output_workbook = Workbook()
output_worksheet = output_workbook.active
output_worksheet.title = 'filtered_rows_all_worksheets'

sales_column_index = 3
threshold = 2000.0

first_worksheet = True
workbook = load_workbook(input_file)
for worksheet in workbook.worksheets:
    if first_worksheet:
        header_row = [cell.value for cell in worksheet[1]]
        output_worksheet.append(header_row)
        first_worksheet = False
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        sale_amount = row[sales_column_index]  # Adjust index to 0-based
        sale_amount = float(str(sale_amount).replace('$', '').replace(',', ''))
        if sale_amount > threshold:
            output_row = []
            for cell_value in row:
                if isinstance(cell_value, date):
                    cell_value = cell_value.strftime('%m/%d/%Y')
                output_row.append(cell_value)
            output_worksheet.append(output_row)

workbook.close()
output_workbook.save(output_file)
