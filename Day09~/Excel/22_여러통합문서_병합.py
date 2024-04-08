#!/usr/bin/env python3
import glob
import os
import sys
from datetime import date
from openpyxl import Workbook, load_workbook

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력폴더, 출력파일
input_folder = path + '/input/'
output_file = path + '/output/' + input('출력 파일 : ')

output_workbook = Workbook()
output_worksheet = output_workbook.active
output_worksheet.title = 'all_data_all_workbooks'

data = []
first_worksheet = True
for input_file in glob.glob(os.path.join(input_folder, '*.xls*')):
    print(os.path.basename(input_file))
    workbook = load_workbook(filename=input_file, read_only=True)
    for worksheet in workbook:
        if first_worksheet:
            header_row = [cell.value for cell in worksheet[1]]
            data.append(header_row)
            first_worksheet = False
        for row in worksheet.iter_rows(min_row=2):
            row_list = []
            for cell in row:
                cell_value = cell.value
                if isinstance(cell_value, date):
                    cell_value = cell_value.strftime('%m/%d/%Y')
                row_list.append(cell_value)
            data.append(row_list)

for row_index, output_list in enumerate(data, start=1):
    for column_index, element in enumerate(output_list, start=1):
        output_worksheet.cell(row=row_index, column=column_index, value=element)

output_workbook.save(output_file)
