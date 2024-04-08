#!/usr/bin/env python3
from openpyxl import load_workbook

import os

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력파일, 출력파일
input_file = path + '/input/' + input('입력 파일 : ')

workbook = load_workbook(input_file)
print('Number of worksheets:', len(workbook.sheetnames))
for worksheet in workbook:
    print("Worksheet name:", worksheet.title, "\tRows:", worksheet.max_row, "\tColumns:", worksheet.max_column)
