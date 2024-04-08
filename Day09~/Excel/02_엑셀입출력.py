#!/usr/bin/env python3
import os
from openpyxl import Workbook
from openpyxl import load_workbook

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력파일, 출력파일
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

output_workbook = Workbook()
output_worksheet = output_workbook.active
output_worksheet.title = 'jan_2013_output'

workbook = load_workbook(input_file)
worksheet = workbook['january_2013']

for row_index, row in enumerate(worksheet.iter_rows(), 1):
    for column_index, cell in enumerate(row, 1):
        output_worksheet.cell(row=row_index, column=column_index, value=cell.value)

output_workbook.save(output_file)
