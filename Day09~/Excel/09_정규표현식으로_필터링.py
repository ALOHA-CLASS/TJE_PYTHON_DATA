#!/usr/bin/env python3
import sys
import os
import re
from datetime import date
from openpyxl import Workbook, load_workbook

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력파일, 출력파일
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

output_workbook = Workbook()
output_worksheet = output_workbook.active
output_worksheet.title = 'jan_2013_output'

pattern = re.compile(r'(?P<my_pattern>^J.*)')

customer_name_column_index = 1

workbook = load_workbook(input_file, read_only=True)
worksheet = workbook['january_2013']
data = []
header = [cell.value for cell in worksheet[1]]
data.append(header)
for row in worksheet.iter_rows(min_row=2, values_only=True):
    row_list = []
    if pattern.match(row[customer_name_column_index]):
        for cell_value in row:
            if isinstance(cell_value, date):
                cell_value = cell_value.strftime('%m/%d/%Y')
            row_list.append(cell_value)
    if row_list:
        data.append(row_list)

for list_index, output_list in enumerate(data):
    for element_index, element in enumerate(output_list):
        output_worksheet.cell(row=list_index+1, column=element_index+1, value=element)

workbook.close()
output_workbook.save(output_file)
