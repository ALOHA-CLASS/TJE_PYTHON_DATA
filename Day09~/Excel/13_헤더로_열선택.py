#!/usr/bin/env python3
import sys
from datetime import date
from openpyxl import Workbook, load_workbook

import os

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력파일, 출력파일
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

output_workbook = Workbook()
output_worksheet = output_workbook.active
output_worksheet.title = 'jan_2013_output'

my_columns = ['Customer ID', 'Purchase Date']

workbook = load_workbook(input_file, read_only=True)
worksheet = workbook['january_2013']
data = [my_columns]
header_list = worksheet[1]
header_index_list = []
for header_index in range(len(header_list)):
    if header_list[header_index].value in my_columns:
        header_index_list.append(header_index)
for row_index in range(2, worksheet.max_row + 1):
    row_list = []
    for column_index in header_index_list:
        cell_value = worksheet.cell(row=row_index, column=column_index + 1).value
        if isinstance(cell_value, date):
            cell_value = cell_value.strftime('%m/%d/%Y')
        row_list.append(cell_value)
    data.append(row_list)

for list_index, output_list in enumerate(data):
    for element_index, element in enumerate(output_list):
        output_worksheet.cell(row=list_index + 1, column=element_index + 1, value=element)

workbook.close()
output_workbook.save(output_file)
