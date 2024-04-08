#!/usr/bin/env python3
import glob
import os
import sys
from openpyxl import load_workbook

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력경로
input_directory = path + '/input/'


workbook_counter = 0
for input_file in glob.glob(os.path.join(input_directory, '*.xlsx')):
    workbook = load_workbook(filename=input_file, read_only=True)
    print('Workbook: {}'.format(os.path.basename(input_file)))
    print('Number of worksheets: {}'.format(len(workbook.sheetnames)))
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        print('Worksheet name:', sheet_name, '\tRows:',\
                worksheet.max_row, '\tColumns:', worksheet.max_column)
    workbook_counter += 1
print('Number of Excel workbooks: {}'.format(workbook_counter))
