#!/usr/bin/env python3
import sys
from datetime import date
from openpyxl import Workbook, load_workbook

import os

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력파일, 출력파일
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

output_workbook = Workbook()
output_worksheet = output_workbook.active
output_worksheet.title = 'selected_columns_all_worksheets'

my_columns = ['Customer Name', 'Sale Amount']

first_worksheet = True
workbook = load_workbook(input_file)
for worksheet in workbook.worksheets:
    data = [my_columns]
    index_of_cols_to_keep = []
    if first_worksheet:
        header_row = [cell.value for cell in worksheet[1]]
        for col_index, header_value in enumerate(header_row):
            if header_value in my_columns:
                index_of_cols_to_keep.append(col_index)
        first_worksheet = False
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_list = []
        for col_index in index_of_cols_to_keep:
            cell_value = row[col_index]
            if isinstance(cell_value, date):
                cell_value = cell_value.strftime('%m/%d/%Y')
            row_list.append(cell_value)
        data.append(row_list)

    for list_index, output_list in enumerate(data):
        for element_index, element in enumerate(output_list):
            output_worksheet.cell(row=list_index + 1, column=element_index + 1, value=element)

workbook.close()
output_workbook.save(output_file)
