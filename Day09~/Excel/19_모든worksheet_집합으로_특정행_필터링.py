#!/usr/bin/env python3
import sys
from datetime import date
from openpyxl import Workbook, load_workbook

import os

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력파일, 출력파일
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

output_workbook = Workbook()
output_worksheet = output_workbook.active
output_worksheet.title = 'set_of_worksheets'

my_sheets = [0, 1]
threshold = 1900.0
sales_column_index = 3

first_worksheet = True
workbook = load_workbook(input_file)
for sheet_index in range(len(workbook.sheetnames)):
    if sheet_index in my_sheets:
        worksheet = workbook[workbook.sheetnames[sheet_index]]
        data = []
        if first_worksheet:
            header_row = [cell.value for cell in worksheet[1]]
            data.append(header_row)
            first_worksheet = False
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_list = []
            sale_amount = row[sales_column_index]
            if sale_amount > threshold:
                for cell_value in row:
                    if isinstance(cell_value, date):
                        cell_value = cell_value.strftime('%m/%d/%Y')
                    row_list.append(cell_value)
            if row_list:
                data.append(row_list)

        for list_index, output_list in enumerate(data):
            for element_index, element in enumerate(output_list):
                output_worksheet.cell(row=list_index + 1, column=element_index + 1, value=element)

workbook.close()
output_workbook.save(output_file)
