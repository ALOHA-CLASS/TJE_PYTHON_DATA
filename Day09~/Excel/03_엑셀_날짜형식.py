#!/usr/bin/env python3
import sys
import os
from datetime import datetime
import openpyxl as px

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력파일, 출력파일
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

output_workbook = px.Workbook()
output_worksheet = output_workbook.active
output_worksheet.title = 'jan_2013_output'

workbook = px.load_workbook(input_file, read_only=True)
worksheet = workbook['january_2013']
for row_index, row in enumerate(worksheet.iter_rows(), 1):
    for col_index, cell in enumerate(row, 1):
        if cell.data_type == 'n' and cell.number_format == 'mm/dd/yyyy':
            date_cell = cell.value.strftime('%m/%d/%Y')
            output_worksheet.cell(row=row_index, column=col_index, value=date_cell)
        else:
            output_worksheet.cell(row=row_index, column=col_index, value=cell.value)

output_workbook.save(output_file)
