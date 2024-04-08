#!/usr/bin/env python3
import sys
import os
from datetime import datetime
import openpyxl as px

# 실행 프로그램의 경로
program_path = os.path.abspath(__file__)
# 디렉터리 경로 - 이 안의 input, output 폴더에서 입출력한다.
path = os.path.dirname(program_path)
# 입력파일, 출력파일
input_file = path + '/input/' + input('입력 파일 : ')
output_file = path + '/output/' + input('출력 파일 : ')

output_workbook = px.Workbook()
output_worksheet = output_workbook.active
output_worksheet.title = 'jan_2013_output'

important_dates = ['01/24/2013', '01/31/2013']

purchase_date_column_index = 5

workbook = px.load_workbook(input_file, read_only=True)
worksheet = workbook['january_2013']

data = []
header = [cell.value for cell in worksheet[1]]
data.append(header)

for row in worksheet.iter_rows(min_row=2, values_only=True):
    purchase_date = row[purchase_date_column_index - 1]
    if purchase_date.strftime('%m/%d/%Y') in important_dates:
        data.append(list(row))

for row_index, output_list in enumerate(data, 1):
    for col_index, value in enumerate(output_list, 1):
        output_worksheet.cell(row=row_index, column=col_index, value=value)

output_workbook.save(output_file)
