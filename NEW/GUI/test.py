from openpyxl import load_workbook      # 엑셀 입력
from openpyxl import Workbook           # 엑셀 출력         💛 추가

import tkinter as tk
from tkinter import filedialog


def open_input_file_dialog():
    filename = filedialog.askopenfilename()
    if filename:
        input_file.set(filename)

def open_output_file_dialog():
    filename = filedialog.asksaveasfilename()
    if filename:
        output_file.set(filename)

def on_button_click():
    label.config(text="입력 파일: {}\n출력 파일: {}".format(input_file.get(), output_file.get()))
    run()

def run():
    # 엑셀 통합 문서 열기 (입력)
    workbook = load_workbook(input_file.get())
    # january_2013 워크시트만 입력
    worksheet = workbook['january_2013']

    # 엑셀 출력 객체 생성
    output_workbook = Workbook()
    output_worksheet =  output_workbook.active  # 워크시트 활성화
    output_worksheet.title = 'out_january_2013' # 워크시트 이름 지정

    # sales_2013.xlsx 의 january_2013 워크시트를 반복하여
    # output02.xlsx 의 out_january_2013 워키스트로 출력
    # 행 반복
    for row_index, row in  enumerate( worksheet.iter_rows(), 1 ):
        # 열 반복
        for column_index, cell in enumerate(row, 1 ):
            # output_worksheet.cell(row=행, column=열, value=값) : 셀의 값을 지정
            output_worksheet.cell(row=row_index, column=column_index, value=cell.value)


    # 엑셀 통합 문서 저장
    output_workbook.save(output_file.get())


# 윈도우 생성
window = tk.Tk()
window.title("파일 선택 예제")

# 창 크기 지정
window.geometry("600x400")

# 라벨 생성
label = tk.Label(window, text="파일을 선택하세요.", padx=20, pady=10)
label.pack()

# 입력 파일 선택 버튼 생성
input_button = tk.Button(window, text="입력 파일 선택", command=open_input_file_dialog)
input_button.pack()

# 선택된 입력 파일 경로를 보여줄 입력 상자 생성
input_file = tk.StringVar()
input_entry = tk.Entry(window, textvariable=input_file, width=50)
input_entry.pack()

# 출력 파일 선택 버튼 생성
output_button = tk.Button(window, text="출력 파일 선택", command=open_output_file_dialog)
output_button.pack()

# 선택된 출력 파일 경로를 보여줄 입력 상자 생성
output_file = tk.StringVar()
output_entry = tk.Entry(window, textvariable=output_file, width=50)
output_entry.pack()

# 실행 버튼 생성
button = tk.Button(window, text="실행", command=on_button_click)
button.pack()

# 윈도우 실행
window.mainloop()
